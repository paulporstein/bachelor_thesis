import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication,QDoubleSpinBox,QWidget,QDialog,QVBoxLayout,QFileDialog,QPushButton,QLabel,QCheckBox,QScrollArea,QRadioButton,QFrame,QSlider
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from PyQt5.uic import loadUi
import openpyxl as op
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from mpl_toolkits import mplot3d
import os
import time
import statistics
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d import proj3d
from arrow_3D import *

class MainWindow(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("GUI.ui",self) #UI-File laden
        
        #Widgets definieren, Widgets sind Kindern von des Eltern-Widgets
        self.btn_browse=self.findChild(QPushButton, "btn_browse")
        self.btn_settings=self.findChild(QPushButton, "btn_settings")
        self.btn_start=self.findChild(QPushButton, "btn_start")
        self.label_file=self.findChild(QLabel,"label_file")
        self.radio_urs=self.findChild(QRadioButton,"radio_urs")
        self.radio_kdm=self.findChild(QRadioButton,"radio_kdm")
        self.frame_plot=self.findChild(QFrame,"frame")
        self.slider_plot=self.findChild(QSlider,"slider_plot")
        self.label_logo=self.findChild(QLabel,"label_logo")
        self.label_step=self.findChild(QLabel,"label_step")
        self.scroll_area=self.findChild(QScrollArea,"scroll_area")
        
        #Radio Button KDM standardmäßig abhaken
        self.radio_kdm.setChecked(True)

        #Funktionen zuweisen
        self.btn_browse.clicked.connect(self.get_filename_and_compute_matrices)
        self.btn_start.clicked.connect(self.start_formfinding)
        self.slider_plot.valueChanged.connect(self.slide_plot)
        self.btn_start.setEnabled(False)
        self.btn_settings.clicked.connect(self.open_settings)
        self.radio_urs.clicked.connect(self.label_spin_urs)
        self.radio_kdm.clicked.connect(self.label_spin_kdm)
        
        #Plotting einrichten
        self.horizontal_layout=QtWidgets.QHBoxLayout(self.frame_plot) #Layout dem Frame zugeordnet
        self.horizontal_layout.setObjectName("horizontal_Layout")
        self.figure=plt.figure("main")
        self.canvas=FigureCanvas(self.figure) #macht ein Widget aus der Figure
        self.horizontal_layout.addWidget(self.canvas) #fügt Widget dem Layout hinzu
        self.horizontal_layout.setContentsMargins(5, 5, 5, 5)
        self.canvas.hide()
        
        #Slider zunächst nicht aktivieren, da sonst bei Bewegung Error, da u.U list_steps noch nicht erstellt
        self.slider_plot.setEnabled(False)
        self.slider_plot.setValue(0)
     
    """
    Label an den Reglern ändern wenn URS oder KDM ausgewählt
    """
     
    def label_spin_urs(self): #wenn noch keine Daten eingelsen wurde, dann wird es bei Klicken der Radio-Buttons eine Fehlermeldung geben
        for j in range (0,np.shape(self.C_s)[0]):
            self.spin_box[j].setPrefix("N{} = ".format(settings.get_sub(str(j+1))))
        
    def label_spin_kdm(self):
        for j in range (0,np.shape(self.C_s)[0]):
            self.spin_box[j].setPrefix("q{} = ".format(settings.get_sub(str(j+1))))
        
    
    """
    Daten einlesen
    """
        
    def get_filename_and_compute_matrices(self):
        filename,_=QFileDialog.getOpenFileName(self,"","","Excel Datei (*.xlsx)")
        self.label_file.setText(os.path.basename(filename))        
        file=op.load_workbook(str(filename))
        #gloabe Varibalen festlegen, damit auch in anderen Funktionen verwendbar
        #global x,y,z,f_x_all,f_y_all,f_z_all,list_fix,list_var,vorgabe,C_s
        #Knoten einlesen
        knoten_sheet=file["Knoten"]
        rows_knoten=knoten_sheet.max_row
        x=np.zeros([rows_knoten-1,1])
        y=np.zeros([rows_knoten-1,1])
        z=np.zeros([rows_knoten-1,1])
        f_x_all=np.zeros([rows_knoten-1,1])
        f_y_all=np.zeros([rows_knoten-1,1])
        f_z_all=np.zeros([rows_knoten-1,1])
        list_fix=list()
        for i in range(2,rows_knoten+1):
            x[i-2]=knoten_sheet.cell(i,2).value
        for i in range(2,rows_knoten+1):
            y[i-2]=knoten_sheet.cell(i,3).value
        for i in range(2,rows_knoten+1):
            z[i-2]=knoten_sheet.cell(i,4).value
        for i in range(2,rows_knoten+1):
            f_x_all[i-2]=knoten_sheet.cell(i,5).value
        for i in range(2,rows_knoten+1):
            f_y_all[i-2]=knoten_sheet.cell(i,6).value
        for i in range(2,rows_knoten+1):
            f_z_all[i-2]=knoten_sheet.cell(i,7).value
        for i in range(2,rows_knoten+1):
            if knoten_sheet.cell(i,8).value == "j":
                list_fix.append(knoten_sheet.cell(i,1).value)
        list_fix=np.array(list_fix)-1
        list_var=np.arange(rows_knoten-1)
        for i in range(0,len(list_fix)):
            for j in range(0,len(list_var)):
                if list_var[j]==list_fix[i]:
                    list_var[j]=0        
        list_var=list_var[list_var!=0]
        #Kanten einlesen
        kanten_sheet=file["Kanten"]
        rows_kanten=kanten_sheet.max_row
        vorgabe=np.zeros([rows_kanten-1])
        C_s=np.zeros([rows_kanten-1,rows_knoten-1])
        for i in range (2,rows_kanten+1):
            vorgabe[i-2]=kanten_sheet.cell(i,4).value
        for i in range (2,rows_kanten+1):
            C_s[i-2,kanten_sheet.cell(i,2).value-1]=1
        for i in range (2,rows_kanten+1):
            C_s[i-2,kanten_sheet.cell(i,3).value-1]=-1
        file.close()
        #Eingabeparameter als Attribute speichern, um überall darauf zugreifen zu können, globale Attribute festzulegen wäre auch möglich
        self.x=x
        self.y=y
        self.z=z
        self.f_x_all=f_x_all
        self.f_y_all=f_y_all
        self.f_z_all=f_z_all
        self.list_fix=list_fix
        self.list_var=list_var
        self.vorgabe=vorgabe
        self.C_s=C_s
        #Formfindung starten Button aktivieren
        self.btn_start.setEnabled(True)
        #ScrollArea anlegen
        self.w=QWidget(self.scroll_area) #ScrollArea als parent von unserem großen
        layout_scroll=QVBoxLayout() #Layout für das große
        self.spin_box=list()
        for j in range (0,np.shape(self.C_s)[0]):
            spin_q=QDoubleSpinBox()
            spin_q.setValue(self.vorgabe[j])
            spin_q.setMaximum(5*self.vorgabe[j])
            if self.radio_kdm.isChecked()==True:
                spin_q.setPrefix("q{} = ".format(settings.get_sub(str(j+1))))
            else:
                spin_q.setPrefix("N{} = ".format(settings.get_sub(str(j+1))))
            self.spin_box.append(spin_q)
            layout_scroll.addWidget(self.spin_box[j]) #Widgets werden dem "großen" Layout hinzugefügt
        self.w.setLayout(layout_scroll) #großes Layout wird dem großen Widget hinzugeügt
        self.scroll_area.setWidget(self.w) #Widget wird der ScrollArea zugewiesen

    """
    Formfindung
    """

    def start_formfinding(self):
        #Daten von den Reglern für q bzw. N abfragen, falls was geändert
        for j in range (0,np.shape(self.C_s)[0]):
            self.vorgabe[j]=self.spin_box[j].value()
        #jetzt kanns los gehen
        self.canvas.show()
        self.label_logo.hide()
        self.list_steps=list()
        if self.radio_kdm.isChecked()==True:
            widget_info.hide()
            self.list_steps.append([self.x,self.y,self.z,0])
            x_kdm,y_kdm,z_kdm,q_kdm=self.kdm(self.x,self.y,self.z,np.diag(self.vorgabe))
            self.list_steps.append([x_kdm,y_kdm,z_kdm,q_kdm])
        if self.radio_urs.isChecked()==True:
            self.list_steps.append([self.x,self.y,self.z,0]) #es wird immer das q angeängt, das für die Ermittlung der Koordinaten an gleicher Position in self.list_steps nötig war
            time_start=time.time()
            self.urs(self.list_steps)
            time_end=time.time()
            if settings.check_information_urs.isChecked()==True:
                info.label_iterationsschritte.setText("Iterationsschritte: {}".format(len(self.list_steps)-1))
                info.label_genauigkeit.setText(("Genauigkeit: max (N{} - N{}) < d{} = 10{}").format(settings.get_sub("j"), settings.get_sub("max,j"),settings.get_sub("max"),settings.get_super(str(settings.slider_d_max.value()))))
                info.label_rechenzeit.setText("Benötigte Zeit: {}s".format(round(time_end-time_start,4)))
                info.slider_plot_nodes.setMaximum(np.shape(self.C_s)[1]-1)
                info.slider_plot_nodes.setValue(0)
                info.plot_coords(self.list_steps,0)
                info.slider_plot_members.setMaximum(np.shape(self.C_s)[0]-1)
                info.slider_plot_members.setValue(0)
                info.plot_members(self.list_steps,0)  
                widget_info.show()
            else:
                widget_info.hide()
        self.slider_plot.setMaximum(len(self.list_steps)-1)
        self.slider_plot.setValue(len(self.list_steps)-1)
        self.plot(self.list_steps[len(self.list_steps)-1][0],self.list_steps[len(self.list_steps)-1][1],self.list_steps[len(self.list_steps)-1][2],self.list_steps[len(self.list_steps)-1][3]) #Endgeometrie soll erstmal geplottet werden
        self.slider_plot.setEnabled(True)
        

    def kdm(self,x_start,y_start,z_start,Q):
        x_kdm=np.array(x_start) #muss nochmal np.array() machen, da aus irgendeinem Grund sonst auch self.x und x_start geändert werden
        y_kdm=np.array(y_start)
        z_kdm=np.array(z_start)
        D_alle=np.transpose(self.C_s)@Q@self.C_s
        D=np.delete(np.delete(D_alle,self.list_fix,0),self.list_fix,1)
        D_f=np.zeros((D.shape[0],len(self.list_fix)))
        for i in range(0,len(self.list_fix)):
            D_f[:,i]=np.delete(D_alle,self.list_fix,0)[:,self.list_fix[i]]
        x_f=np.zeros([len(self.list_fix),1])
        y_f=np.zeros([len(self.list_fix),1])
        z_f=np.zeros([len(self.list_fix),1])
        for i in range (0,len(self.list_fix)):
            x_f[i]=x_start[self.list_fix[i]]
            y_f[i]=y_start[self.list_fix[i]]
            z_f[i]=z_start[self.list_fix[i]]
        f_x=np.delete(self.f_x_all,self.list_fix,0)
        f_y=np.delete(self.f_y_all,self.list_fix,0)
        f_z=np.delete(self.f_z_all,self.list_fix,0)
        x_var=np.linalg.solve(D,f_x-D_f@x_f)
        y_var=np.linalg.solve(D,f_y-D_f@y_f)
        z_var=np.linalg.solve(D,f_z-D_f@z_f)
        for i in range(0,len(self.list_var)):
            x_kdm[self.list_var[i]]=x_var[i] #kann man so machen, da Punkte in richter Reihnfolge
            y_kdm[self.list_var[i]]=y_var[i]
            z_kdm[self.list_var[i]]=z_var[i]
        return x_kdm,y_kdm,z_kdm,np.diag(Q)
        
    def urs(self,list_steps):
        abfrage=True
        x_start=np.array(list_steps[0][0])
        y_start=np.array(list_steps[0][1])
        z_start=np.array(list_steps[0][2])
        i=0
        r=np.square(self.C_s@x_start)
        s=np.square(self.C_s@y_start)
        t=np.square(self.C_s@z_start)
        l=np.sqrt(r+s+t) #Länge aller Stäbe für die Startgeometrie
        while abfrage==True:
            q=self.vorgabe/l.flatten()
            Q=np.diag(q)
            x_kdm,y_kdm,z_kdm,q_kdm=self.kdm(list_steps[i][0],list_steps[i][1],list_steps[i][2],Q)
            list_steps.append([x_kdm,y_kdm,z_kdm,q_kdm])
            r_neu=np.square(self.C_s@list_steps[i+1][0])
            s_neu=np.square(self.C_s@list_steps[i+1][1])
            t_neu=np.square(self.C_s@list_steps[i+1][2])
            l=np.sqrt(r_neu+s_neu+t_neu) #Länge der im (n-1)-ten Iterationsschritts gefundenen Stabelemente = Referenzgeometrie für nächsten Schritt
            n_neu=q_kdm*l.flatten()
            max_diff=max(abs(n_neu-self.vorgabe))
            if max_diff<settings.d_max:
                abfrage=False
            else:
                i=i+1
            if i== settings.slider_steps_abbruch.value(): #Abbrechen, wenn Iteration zu lange dauert wegen z.B. Inkompatibität des vorgegebenen Spannungszustands
                abfrage=False
                abbruch.label_hinweis_abbruch.setText("Iteration wurde automatisch nach "+str(settings.slider_steps_abbruch.value())+" Schritten beendet!")
                widget_abbruch.show()

    """
    Plotten
    """
    
    def set_axes_equal(self,ax: plt.Axes):
        """
        Set 3D plot axes to equal scale.    
        Make axes of 3D plot have equal scale so that spheres appear as
        spheres and cubes as cubes.  Required since `ax.axis('equal')`
        and `ax.set_aspect('equal')` don't work on 3D.
        """
        limits = np.array([
            ax.get_xlim3d(),
            ax.get_ylim3d(),
            ax.get_zlim3d(),
            ])
        origin = np.mean(limits, axis=1)
        radius = 0.5 * np.max(np.abs(limits[:, 1] - limits[:, 0]))
        self._set_axes_radius(ax, origin, radius)
        
    def _set_axes_radius(self,ax, origin, radius):
        x, y, z = origin
        ax.set_xlim3d([x - radius, x + radius])
        ax.set_ylim3d([y - radius, y + radius])
        ax.set_zlim3d([z - radius, z + radius])
    
    def slide_plot(self,value):
        self.plot(self.list_steps[value][0],self.list_steps[value][1],self.list_steps[value][2],self.list_steps[value][3])
        if value==0:
            self.label_step.setText("Ausgangsgeometrie")
        else:
            self.label_step.setText("Iteration"+" "+str(value))

    def plot(self,x,y,z,q):
        plt.figure("main")
        #bring in flat shape to plot, nur lokale Variablen
        x=x.flatten()
        y=y.flatten()
        z=z.flatten()
        plt.clf()
        ax = plt.axes(projection ="3d")
        ax.clear()
        ax.scatter3D(x, y, z,s=5)
        if settings.check_axes_equal.isChecked()==True:
            self.set_axes_equal(ax)
        for j in range(0,self.C_s.shape[0]):
            for i in range(0,self.C_s.shape[1]):
                if self.C_s[j,i]==1:
                    x_start=x[i]
                    y_start=y[i]
                    z_start=z[i]
                if self.C_s[j,i]==-1:
                    x_end=x[i]
                    y_end=y[i]
                    z_end=z[i]
            ax.plot3D((x_start,x_end),(y_start,y_end),(z_start,z_end),color="grey",linewidth=1)
        if settings.check_annotate_nodes.isChecked()==True:
            for i in range (0,len(x)):
                ax.text(x[i],y[i],z[i],"i="+str(i+1),color="blue")
        if settings.check_annotate_members.isChecked()==True:
            for j in range (0,np.shape(self.C_s)[0]):
                x_points=list()
                y_points=list()
                z_points=list()
                for i in range (0,np.shape(self.C_s)[1]):
                    if self.C_s[j,i]!=0:
                        x_points.append(x[i])
                        y_points.append(y[i])
                        z_points.append(z[i])
                x_middle=statistics.median(x_points)
                y_middle=statistics.median(y_points)
                z_middle=statistics.median(z_points)
                ax.text(x_middle,y_middle,z_middle,"j="+str(j+1),color="black")
        if settings.check_annotate_q.isChecked()==True:
            for j in range (0,np.shape(self.C_s)[0]):
                x_points=list()
                y_points=list()
                z_points=list()
                for i in range (0,np.shape(self.C_s)[1]):
                    if self.C_s[j,i]!=0:
                        x_points.append(x[i])
                        y_points.append(y[i])
                        z_points.append(z[i])
                x_middle=statistics.median(x_points)
                y_middle=statistics.median(y_points)
                z_middle=statistics.median(z_points)
                if self.slider_plot.value()!=0:
                    ax.text(x_middle,y_middle,z_middle,"q{}=".format(settings.get_sub(str(j+1)))+str(round(q[j],2)),color="purple")
        if settings.check_forces.isChecked()==True:
            for i in range (0,np.shape(self.C_s)[1]):
                if self.f_x_all[i]!=0:
                    if self.f_x_all[i] <0:
                        abstand=-(abs((max(x)-min(x))/4))
                    else:
                        abstand=abs((max(x)-min(x))/4)
                    ax.arrow3D(x[i],y[i],z[i],abstand,0,0,mutation_scale=10,arrowstyle="-|>",color="red")            
                    ax.text(x[i]+1.1*abstand,y[i],z[i],"$f_{x}=$"+str(abs(float(self.f_x_all[i]))),color="red")
                if self.f_y_all[i]!=0:
                    if self.f_y_all[i] <0:
                        abstand=-(abs((max(y)-min(y))/4))
                    else:
                        abstand=abs((max(y)-min(y))/4)
                    ax.arrow3D(x[i],y[i],z[i],0,abstand,0,mutation_scale=10,arrowstyle="-|>",color="red")            
                    ax.text(x[i],y[i]+1.1*abstand,z[i],"$f_{y}=$"+str(abs(float(self.f_y_all[i]))),color="red")
                if self.f_z_all[i]!=0:
                    if self.f_z_all[i] <0:
                        abstand=-(abs((max(z)-min(z))/4))
                    else:
                        abstand=abs((max(z)-min(z))/4)
                    ax.arrow3D(x[i],y[i],z[i],0,0,abstand,mutation_scale=10,arrowstyle="-|>",color="red")            
                    ax.text(x[i],y[i],z[i]+1.1*abstand,"$f_{z}=$"+str(abs(float(self.f_z_all[i]))),color="red")
        if settings.check_annotate_alr.isChecked()==True:
            f=np.transpose(self.C_s)@np.diag(q)@self.C_s@x-self.f_x_all.flatten()
            g=np.transpose(self.C_s)@np.diag(q)@self.C_s@y-self.f_y_all.flatten()
            h=np.transpose(self.C_s)@np.diag(q)@self.C_s@z-self.f_z_all.flatten()
            for i in range (0,len(f)):
                if i in self.list_fix:
                    if f[i] <0:
                        abstand=-(abs((max(x)-min(x))/4))
                    else:
                        abstand=abs((max(x)-min(x))/4)
                    ax.arrow3D(x[i],y[i],z[i],abstand,0,0,mutation_scale=10,arrowstyle="-|>",color="green")            
                    ax.text(x[i]+1.1*abstand,y[i],z[i],"$A_{x}=$"+str(abs(round((float(f[i])),2))),color="green")
            for i in range (0,len(g)):
                if i in self.list_fix:
                    if g[i] <0:
                        abstand=-(abs((max(y)-min(y))/4))
                    else:
                        abstand=abs((max(y)-min(y))/4)
                    ax.arrow3D(x[i],y[i],z[i],0,abstand,0,mutation_scale=10,arrowstyle="-|>",color="green")            
                    ax.text(x[i],y[i]+1.1*abstand,z[i],"$A_{y}=$"+str(abs(round((float(g[i])),2))),color="green")
            for i in range (0,len(h)):
                if i in self.list_fix:
                    if h[i] <0:
                        abstand=-(abs((max(z)-min(z))/4))
                    else:
                        abstand=abs((max(z)-min(z))/4)
                    ax.arrow3D(x[i],y[i],z[i],0,0,abstand,mutation_scale=10,arrowstyle="-|>",color="green")            
                    ax.text(x[i],y[i],z[i]+1.1*abstand,"$A_{z}=$"+str(abs(round((float(h[i])),2))),color="green")
        #ax.set_axis_off()
        #plt.savefig("Iteration_"+str(self.slider_plot.value())+".png",dpi=500,bbox_inches='tight')
        #ax.set_axis_on()
        self.canvas.draw()
        
        
    """
    Settings
    """
    
    def open_settings(self):
        widget_settings.show()
    
class Settings(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("GUI_Settings.ui",self) #UI-File laden mit  geladener Funktion
        
        #Widgets definieren
        self.btn_reset=self.findChild(QPushButton, "btn_reset")
        self.label_d_max=self.findChild(QLabel, "label_d_max")
        self.slider_d_max=self.findChild(QSlider, "slider_d_max")
        self.check_information_urs=self.findChild(QCheckBox, "check_information_urs")
        self.check_annotate_members=self.findChild(QCheckBox,"check_annotate_members")
        self.check_annotate_nodes=self.findChild(QCheckBox,"check_annotate_nodes")
        self.check_annotate_q=self.findChild(QCheckBox,"check_annotate_q")
        self.check_forces=self.findChild(QCheckBox,"check_forces")
        self.check_annoate_alr=self.findChild(QCheckBox,"check_annotate_alr")
        self.check_axes_equal=self.findChild(QCheckBox,"check_axes_equal")
        self.label_steps_abbruch=self.findChild(QLabel, "label_steps_abbruch")
        self.slider_steps_abbruch=self.findChild(QSlider, "slider_steps_abbruch")



        #Funktionen zuweisen
        self.btn_reset.clicked.connect(self.reset)
        self.slider_d_max.valueChanged.connect(self.change_d_max)
        self.slider_steps_abbruch.valueChanged.connect(self.change_steps_abbruch)
        
        #Erklärungen für checkBoxes anlegen
        self.check_annotate_q.setToolTip("Plottet die Kraftdichten, mit denen die aktuell gezeigte Geometrie ermittelt wurde")
        self.check_forces.setToolTip("Plottet positiv angetragenen äußeren Lasten")
        self.check_annoate_alr.setToolTip("Plottet positiv angetragenen Auflagerreaktionen")

        
        #d_max einrichten
        a=self.get_sub("max")
        b=self.get_super(str(self.slider_d_max.value()))
        self.label_d_max.setText(("d{} = 10{}").format(a, b))
        self.d_max=pow(10,self.slider_d_max.value())
        self.slider_d_max.setValue(0)
        
        #steps Abbruch einrichten
        self.slider_steps_abbruch.setValue(100)
        
    """
    autoamtischen Abbruch einstellen
    """
    
    def change_steps_abbruch(self,value):
        self.label_steps_abbruch.setText("automatischer Abbruch nach "+str(value)+" Schritten")
        
    """
    Reset
    """
    
    def reset(self):
        mainwindow.slider_plot.setEnabled(False)
        mainwindow.btn_start.setEnabled(False)
        mainwindow.label_logo.show()
        mainwindow.slider_plot.setValue(0)
        mainwindow.canvas.hide()
        mainwindow.label_file.setText("")
        mainwindow.label_step.setText("")
        self.slider_d_max.setValue(0)
        self.label_d_max.setText(("d{} = 10{}").format(self.get_sub("max"), self.get_super(str(self.slider_d_max.value()))))
        self.d_max=pow(10,self.slider_d_max.value())
        widget_info.hide()
        mainwindow.w.hide()
        mainwindow.radio_kdm.setChecked(True)

        
    """
    d_max
    """
    
    def get_super(self,x):
        normal = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-=()"
        super_s = "ᴬᴮᶜᴰᴱᶠᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾQᴿˢᵀᵁⱽᵂˣʸᶻᵃᵇᶜᵈᵉᶠᵍʰᶦʲᵏˡᵐⁿᵒᵖ۹ʳˢᵗᵘᵛʷˣʸᶻ⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾"
        res = x.maketrans(''.join(normal), ''.join(super_s))
        return x.translate(res)
    
    def get_sub(self,x):
        normal = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-=()"
        sub_s = "ₐ₈CDₑբGₕᵢⱼₖₗₘₙₒₚQᵣₛₜᵤᵥwₓᵧZₐ♭꜀ᑯₑբ₉ₕᵢⱼₖₗₘₙₒₚ૧ᵣₛₜᵤᵥwₓᵧ₂₀₁₂₃₄₅₆₇₈₉₊₋₌₍₎"
        res = x.maketrans(''.join(normal), ''.join(sub_s))
        return x.translate(res)
    
    
    def change_d_max(self,value):
        self.d_max=pow(10,value)
        a=self.get_sub("max")
        b=self.get_super(str(value))
        self.label_d_max.setText(("d{} = 10{}").format(a, b))
        self.d_max=pow(10,value)
        
    
class information_urs(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("GUI_Info.ui",self) #UI-File laden mit  geladener Funktion
        
        #Widgets definieren
        self.label_genauigkeit=self.findChild(QLabel, "label_genauigkeit")
        self.label_iterationsschritte=self.findChild(QLabel, "label_iterationsschritte")
        self.label_rechenzeit=self.findChild(QLabel, "label_rechenzeit")
        self.frame_plot_x=self.findChild(QFrame,"frame_plot_x")
        self.frame_plot_y=self.findChild(QFrame,"frame_plot_y")
        self.frame_plot_z=self.findChild(QFrame,"frame_plot_z")
        self.slider_plot_nodes=self.findChild(QSlider,"slider_plot_nodes")
        self.label_info_node=self.findChild(QLabel, "label_info_node")
        self.frame_plot_q=self.findChild(QFrame,"frame_plot_q")
        self.frame_plot_diff_soll=self.findChild(QFrame,"frame_plot_diff_soll")
        self.frame_plot_diff=self.findChild(QFrame,"frame_plot_diff")
        self.slider_plot_members=self.findChild(QSlider,"slider_plot_members")
        self.label_info_member=self.findChild(QLabel, "label_info_member")
        
        #Funktionen zuweisen
        self.slider_plot_nodes.valueChanged.connect(self.slide_plot_coords)
        self.slider_plot_members.valueChanged.connect(self.slide_plot_members)
        
        #Slider und Labels für Plotting einrichten
        self.label_info_node.setText("Knoten {}".format(1))
        self.label_info_member.setText("Stab {}".format(1))
                
        #Plotting einrichten
        self.horizontal_layout_x=QtWidgets.QHBoxLayout(self.frame_plot_x)
        self.horizontal_layout_x.setObjectName("horizontal_Layout")
        self.figure_x=plt.figure("steps_x")
        self.canvas_x=FigureCanvas(self.figure_x)
        self.horizontal_layout_x.addWidget(self.canvas_x)
        self.horizontal_layout_x.setContentsMargins(10, 10, 10, 10)
        
        self.horizontal_layout_y=QtWidgets.QHBoxLayout(self.frame_plot_y)
        self.horizontal_layout_y.setObjectName("horizontal_Layout")
        self.figure_y=plt.figure("steps_y")
        self.canvas_y=FigureCanvas(self.figure_y)
        self.horizontal_layout_y.addWidget(self.canvas_y)
        self.horizontal_layout_y.setContentsMargins(10, 10, 10, 10)
        
        self.horizontal_layout_z=QtWidgets.QHBoxLayout(self.frame_plot_z)
        self.horizontal_layout_z.setObjectName("horizontal_Layout")
        self.figure_z=plt.figure("steps_z")
        self.canvas_z=FigureCanvas(self.figure_z)
        self.horizontal_layout_z.addWidget(self.canvas_z)
        self.horizontal_layout_z.setContentsMargins(10, 10, 10, 10)
        
        self.horizontal_layout_plot_q=QtWidgets.QHBoxLayout(self.frame_plot_q)
        self.horizontal_layout_plot_q.setObjectName("horizontal_Layout")
        self.figure_plot_q=plt.figure("steps_plot_q")
        self.canvas_plot_q=FigureCanvas(self.figure_plot_q)
        self.horizontal_layout_plot_q.addWidget(self.canvas_plot_q)
        self.horizontal_layout_plot_q.setContentsMargins(10, 10, 10, 10)
        
        self.horizontal_layout_plot_diff_soll=QtWidgets.QHBoxLayout(self.frame_plot_diff_soll)
        self.horizontal_layout_plot_diff_soll.setObjectName("horizontal_Layout")
        self.figure_plot_diff_soll=plt.figure("steps_plot_diff_soll")
        self.canvas_plot_diff_soll=FigureCanvas(self.figure_plot_diff_soll)
        self.horizontal_layout_plot_diff_soll.addWidget(self.canvas_plot_diff_soll)
        self.horizontal_layout_plot_diff_soll.setContentsMargins(10, 10, 10, 10)
        
        self.horizontal_layout_plot_diff=QtWidgets.QHBoxLayout(self.frame_plot_diff)
        self.horizontal_layout_plot_diff.setObjectName("horizontal_Layout")
        self.figure_plot_diff=plt.figure("steps_plot_diff")
        self.canvas_plot_diff=FigureCanvas(self.figure_plot_diff)
        self.horizontal_layout_plot_diff.addWidget(self.canvas_plot_diff)
        self.horizontal_layout_plot_diff.setContentsMargins(10, 10, 10, 10)
        
        
    """
    Koordinaten für jeden Iterationsschritt plotten
    """
    
    def slide_plot_coords(self,value):
        self.plot_coords(mainwindow.list_steps,value)
        self.label_info_node.setText("Knoten {}".format(value+1))
    
    def plot_coords(self,list_steps,node):
        x_n=list()
        y_n=list()
        z_n=list()
        plt.figure("steps_x")
        plt.clf()
        plt.cla()
        plt.title("Verlauf x{} in {} Iterationen".format(settings.get_sub(str(node+1)),len(list_steps)-1))
        plt.figure("steps_y")
        plt.clf()
        plt.cla()
        plt.title("Verlauf y{} in {} Iterationen".format(settings.get_sub(str(node+1)),len(list_steps)-1))
        plt.figure("steps_z")
        plt.clf()
        plt.cla()
        plt.title("Verlauf z{} in {} Iterationen".format(settings.get_sub(str(node+1)),len(list_steps)-1))
        for i in range(0,len(list_steps)):
            x_n.append(float(list_steps[i][0][node]))
            y_n.append(float(list_steps[i][1][node]))
            z_n.append(float(list_steps[i][2][node]))
        plt.figure("steps_x")
        plt.plot(np.arange(len(list_steps)),x_n)
        plt.figure("steps_y")
        plt.plot(np.arange(len(list_steps)),y_n)
        plt.figure("steps_z")
        plt.plot(np.arange(len(list_steps)),z_n)
        self.canvas_x.draw()
        self.canvas_y.draw()
        self.canvas_z.draw()
        
        
    """
    Informationen über Stäbe in jedem iterationsschritt plotten
    """
    
    def slide_plot_members(self,value):
        self.plot_members(mainwindow.list_steps,value)
        self.label_info_member.setText("Stab {}".format(value+1))
        
    def plot_members(self,list_steps,member):
        plt.figure("steps_plot_q")
        plt.cla()
        plt.clf()
        plt.title("q{}{} in {} Iterationen".format(settings.get_sub(str(member+1)),settings.get_super("(n)"),len(list_steps)-1))
        plt.figure("steps_plot_diff")
        plt.cla()
        plt.clf()
        plt.title("|N{}{} - N{}{}| in {} Iterationen".format(settings.get_sub(str(member+1)),settings.get_super("(n)"),settings.get_sub(str(member+1)),settings.get_super("(n-1)"),len(list_steps)-1))
        plt.figure("steps_plot_diff_soll")
        plt.cla()
        plt.clf()
        plt.title("|N{}{} - N{}| in {} Iterationen".format(settings.get_sub(str(member+1)),settings.get_super("(n)"),settings.get_sub("soll,"+str(member+1)),len(list_steps)-1))
        q_n=list()
        n_i=list()
        n_diff=list()
        n_diff_soll=list()
        for i in range (0,len(list_steps)):
            if i==0:
                q_n.append(list_steps[i][3])
                n_i.append(list_steps[i][3])
            else:
                q_n.append(list_steps[i][3][member]) #Kraftdichten
                r=np.square(mainwindow.C_s@list_steps[i][0]) #Differenz N zu vorigem Iterationsschritt
                s=np.square(mainwindow.C_s@list_steps[i][1])
                t=np.square(mainwindow.C_s@list_steps[i][2])
                l_i=np.sqrt(r+s+t)
                n_i.append(list_steps[i][3][member]*l_i[member])
        for j in range (1,len(n_i)):
            n_diff.append(n_i[j]-n_i[j-1])
        n_diff_soll=abs(n_i-mainwindow.vorgabe[member])
        plt.figure("steps_plot_q")
        plt.plot(np.arange(len(list_steps)),q_n)
        plt.figure("steps_plot_diff")
        plt.plot(np.arange(len(list_steps)-1)+1,abs(np.array(n_diff)))
        plt.figure("steps_plot_diff_soll")
        plt.plot(np.arange(len(list_steps)),n_diff_soll)        
        self.canvas_plot_q.draw()
        self.canvas_plot_diff.draw()
        self.canvas_plot_diff_soll.draw()
        
    """
    Abbruchfenster für Abbruch der Iteration nach Vorgabe
    """
    
class Abbruch(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("GUI_Abbruch.ui",self)
        
        #Widget definieren
        self.label_hinweis_abbruch=self.findChild(QLabel, "label_hinweis_abbruch")

        
        
    
app=QApplication(sys.argv)

mainwindow=MainWindow()
widget=QtWidgets.QStackedWidget() #QStackedWidgets ist eine "Liste" von QDialogs, hat  jetzt aber  eben nur einen Eintrag
widget.addWidget(mainwindow)
widget.setFixedHeight(500)
widget.setFixedWidth(800)

settings=Settings()
widget_settings=QtWidgets.QStackedWidget()
widget_settings.addWidget(settings)
widget_settings.setFixedHeight(464)
widget_settings.setFixedWidth(280)

info=information_urs()
widget_info=QtWidgets.QStackedWidget()
widget_info.addWidget(info)
widget_info.setFixedHeight(630)
widget_info.setFixedWidth(720)

abbruch=Abbruch()
widget_abbruch=QtWidgets.QStackedWidget()
widget_abbruch.addWidget(abbruch)
widget_abbruch.setFixedHeight(80)
widget_abbruch.setFixedWidth(520)

widget.show()

sys.exit(app.exec())